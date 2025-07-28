import streamlit as st
import pandas as pd
import re
from io import BytesIO
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import numbers
import difflib


st.set_page_config(page_title="CRM Matching Tool", layout="wide")
st.title("🔍 CRM Matching & Classification Tool")

# === Upload files ===
database_file = st.file_uploader("📂 Upload CRM Database File (Excel)", type=["xlsx"])
lookup_file = st.file_uploader("📂 Upload Lookup File (Excel)", type=["xlsx"])
region_file = st.file_uploader("🌍 Upload Country-Region Mapping File (Excel)", type=["xlsx"])


def detect_header_row(file, required_cols, file_label):
    xl = pd.ExcelFile(file)
    sheet_names = xl.sheet_names
    sheet_name = st.selectbox(f"Select sheet from {file_label}:", sheet_names, key=file.name)
    for i in range(10):
        try:
            df_sample = pd.read_excel(file, sheet_name=sheet_name, header=i, nrows=1)
            if all(col in df_sample.columns for col in required_cols):
                st.success(f"✅ Header detected at row {i + 1} in {file_label}")
                return sheet_name, i
        except Exception:
            continue
    return None, None

if database_file and lookup_file:
    required_db_columns = ['Account Name', 'Account Group Name Cleaned', 'Country', 'Business Type']
    required_lookup_columns = ['Company name']



    st.subheader("📋 CRM File Settings")
    db_sheet, db_header = detect_header_row(database_file, required_db_columns, "CRM File")

    st.subheader("📋 Lookup File Settings")
    lookup_sheet, lookup_header = detect_header_row(lookup_file, required_lookup_columns, "Lookup File")

    if db_sheet is None or lookup_sheet is None:
        st.error("❌ Could not detect required headers.")
        st.stop()

    db = pd.read_excel(database_file, sheet_name=db_sheet, header=db_header)

    lookup = pd.read_excel(lookup_file, sheet_name=lookup_sheet, header=lookup_header)
    # If "Company name" column doesn't exist, try to find any likely company name column
    # If "Company name" exists, keep it as is
    if "Company name" in lookup.columns:
        pass  # No renaming needed
    elif "Importer name" in lookup.columns:
        lookup.rename(columns={"Importer name": "Company name"}, inplace=True)
    else:
        # Try to infer a company name column
        potential_names = [col for col in lookup.columns if "name" in col.lower()]
        if potential_names:
            lookup.rename(columns={potential_names[0]: "Company name"}, inplace=True)
        else:
            st.warning("⚠️ No 'Company name' or similar column found. Creating empty 'Company name'.")
            lookup["Company name"] = ""

    # === Rename company and country columns ===
    # === Rename only 'Country' to 'Buying country'; keep 'Company name' unchanged ===
    if "Country" in lookup.columns:
        lookup.rename(columns={"Country": "Buying country"}, inplace=True)

    # === Load region mapping and merge ===
    # === Load region mapping and merge cleanly ===
    if region_file:
        region_df = pd.read_excel(region_file)
        region_df.columns = [col.strip() for col in region_df.columns]

        # Only keep 'Country' and 'Region' columns
        region_df = region_df[['Country', 'Region']].copy()
        region_df.rename(columns={"Country": "Country", "Region": "Buying country Region"}, inplace=True)


        # Normalize function: strip, lowercase, collapse multiple spaces
        def normalize(text):
            return re.sub(r'\s+', ' ', str(text).strip().lower())


        # Create a normalized dictionary
        # Build normalized region dictionary
        country_region_dict = {
            normalize(country): region
            for country, region in zip(region_df['Country'], region_df['Buying country Region'])
        }


        # Matching function with priority rules
        def find_best_match(buying_country):
            if pd.isna(buying_country):
                return ""

            upper_country = str(buying_country).upper()

            # === Priority keyword rules ===
            if "KINGDOM" in upper_country:
                return "West Europe"
            if "OF AM" in upper_country or "UNITED STATES" in upper_country:
                return "North America"
            if "EMIRATES" in upper_country:
                return "Middle East"

            # Normalize for dictionary match
            norm_buying = normalize(buying_country)

            # Fuzzy match (80%)
            matches = difflib.get_close_matches(norm_buying, country_region_dict.keys(), n=1, cutoff=0.8)
            if matches:
                return country_region_dict[matches[0]]

            # Try first 5-letter match
            for norm_country in country_region_dict:
                if norm_buying[:5] == norm_country[:5]:
                    return country_region_dict[norm_country]

            return ""


        # Apply the function
        lookup["Buying country Region"] = lookup["Buying country"].apply(find_best_match)

        # Move the matched column to the right of "Buying country"
        region_col = lookup.pop("Buying country Region")
        insert_pos = lookup.columns.get_loc("Buying country") + 1
        lookup.insert(insert_pos, "Buying country Region", region_col)

    # === Cleaning ===
    def clean_name(name):
        if pd.isna(name): return ''
        name = name.lower()
        name = re.sub(r'[^\w\s]', '', name)
        name = re.sub(r'\b(inc|llc|ltd|co|corporation|company|limited|group|division|plc|gmbh|sa|bv|global|sarl|sro|kg|ltda|operations|applied to life|automotive|packaging)\b', '', name)
        name = re.sub(r'\s+', ' ', name).strip()
        return name

    def get_prefix(name):
        return re.sub(r'\s+', '', name)[:4]

    db['Cleaned Account Name'] = db['Account Name'].apply(clean_name)
    db['Cleaned Group Name'] = db['Account Group Name Cleaned'].apply(clean_name)
    db['Cleaned Name'] = db['Cleaned Account Name']
    db['Prefix'] = db['Cleaned Account Name'].apply(get_prefix)

    lookup['Cleaned Company Name'] = lookup['Company name'].apply(clean_name)
    lookup['Prefix'] = lookup['Cleaned Company Name'].apply(get_prefix)

    def get_threshold(name_len):
        return 88 if name_len <= 12 else 80 if name_len <= 20 else 70

    def match_business_type(cleaned_name, country, prefix):
        if not prefix.strip():
            return "Not in CRM"

        # Allow fallback matching even if prefix not in CRM

        best_score = 0
        best_type = None
        same_country_type = None
        other_region_customer = False
        threshold = get_threshold(len(cleaned_name))

        for _, row in db.iterrows():
            if prefix != row['Prefix']:
                continue
            db_country = str(row['Country']).strip().lower()
            db_type = str(row['Business Type']).strip()
            scores = [
                fuzz.token_set_ratio(cleaned_name, row['Cleaned Account Name']),
                fuzz.token_set_ratio(cleaned_name, row['Cleaned Group Name']),
                fuzz.token_set_ratio(cleaned_name, row['Cleaned Name'])
            ]
            score = max(scores)

            if score >= 90 and (not country or db_country == country):
                return db_type
            if score >= 90 and db_country != country and db_type == "Customer":
                other_region_customer = True
            if score >= threshold and db_country == country and same_country_type is None:
                same_country_type = db_type
            if score > best_score:
                best_score = score
                best_type = db_type

        if same_country_type:
            return same_country_type
        if other_region_customer:
            return "Not in CRM (Customer in other region)"
        if best_score >= (threshold - 5):
            return "Other"
        return "Not in CRM"

    st.write("🔄 Matching business types...")
    lookup['Business Type Matched'] = lookup.apply(
        lambda row: match_business_type(
            row['Cleaned Company Name'],
            str(row['Buying country']).strip().lower() if 'Buying country' in row else "",
            row['Prefix']
        ), axis=1
    )

    # === Classification ===
    # Extract year columns safely and sort numerically
    year_columns = [col for col in lookup.columns if str(col).isdigit()]
    year_columns = sorted(year_columns, key=lambda x: int(x))

    num_years = len(year_columns)
    last_year = year_columns[-1]

    def classify_company(row):
        business_type = str(row.get("Business Type Matched", "")).strip().lower()
        company_name = str(row.get("Company name", "")).strip().lower()
        values = [row.get(y, 0) if pd.notna(row.get(y)) else 0 for y in year_columns]
        grand_total = sum(values)
        last_year_value = row.get(last_year) or 0

        if any(kw in business_type for kw in ["customer", "logistics", "trading"]) or \
           any(kw in company_name for kw in ["logistics", "freight", "trading", "forwarding", "export"]):
            return "D"
        if "prospect" in business_type:
            return "P"
        if last_year_value > 20000:
            if all(v == 0 for v in values[:-1]) or (max(values[:-1]) < 5000 and last_year_value > 30000):
                return "B"
        if last_year_value > 0:
            non_zero = [v for v in values if v > 0]
            if len(non_zero) >= 2 and all(non_zero[i] >= non_zero[i - 1] for i in range(1, len(non_zero))):
                if grand_total >= num_years * 10000 or (last_year_value > 20000 and last_year_value >= 0.6 * grand_total):
                    return "C"
        if grand_total >= num_years * 30000 and last_year_value > 30000:
            return "A"
        if (grand_total >= num_years * 30000 and last_year_value <= 30000) or \
           (last_year_value > 0 and sum(1 for v in values if v > 0) >= 0.8 * num_years and grand_total >= 20000 * num_years):
            return "F"
        return "N"

    st.write("🔠 Classifying companies...")
    lookup['Classification'] = lookup.apply(classify_company, axis=1)
    lookup.drop(columns=["Cleaned Company Name", "Prefix"], inplace=True, errors="ignore")

    # === Save Excel with formatting ===
    st.success("✅ Finished processing!")
    temp_output = BytesIO()
    lookup.to_excel(temp_output, index=False, engine='openpyxl')
    temp_output.seek(0)

    wb = load_workbook(temp_output)
    ws = wb.active

    # Insert classification rules
    description = (
        "Classification (PET Preform ONLY):\n"
        "D = Customer/Logistics/Trading\n"
        "P = Prospect in CRM\n"
        "B = New and sudden growth >20K in latest year\n"
        "C = Increasing trend + (Total >10K/year or (60% in latest year > 20k))\n"
        "A = High-performing (Total >30K/year and latest year >30K)\n"
        "F = Stable or reduced recently (Total > 30k/year and latest year <30k)\n"
        "N = Not interesting to focus"
    )

    ws.insert_rows(1)
    end_col = min(10, ws.max_column)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(row=1, column=1).value = description
    ws.cell(row=1, column=1).fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    ws.cell(row=1, column=1).font = Font(bold=True)

    # Highlight headers
    header_row = 2

    # Collect all numeric-type columns (years and grand total or any number columns)
    numeric_columns = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if header is None:
            continue
        if re.match(r"^\d{4}$", str(header)) or "total" in str(header).lower():
            numeric_columns.append(col_idx)

    # Apply comma-style number format
    for col_idx in numeric_columns:
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # Finalize download
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.download_button(
        label="⬇️ Download Final Excel",
        data=final_output.getvalue(),
        file_name="Processed_Lookup.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
