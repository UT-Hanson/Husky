import streamlit as st
import pandas as pd
import re
import os
import unicodedata
from io import BytesIO
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="🧼 Brand Clustering", layout="wide")
st.title("🔍 Buyer & Supplier Brand Clustering Tool")

# === Cleaning Functions ===
def clean_name(name):
    if pd.isna(name):
        return ""
    name = str(name)
    name = unicodedata.normalize("NFKD", name)
    name = name.encode("ascii", "ignore").decode("utf-8")
    name = name.lower()
    name = re.sub(r'[^a-z0-9 ]', '', name)
    name = re.sub(r'\b(inc|ltd|corp|co|company|limited|group|plc|gmbh|sa|bv|canada|austria|division of .*)\b', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def extract_brand(name):
    words = name.split()
    return words[0] if words else ""

# === Priority brand map ===
priority_brands = {
    'ARBURG': 'ARBURG',
    'SERAC': 'SERAC',
    'KHS': 'KHS',
    'KRONES': 'KRONES',
    'SHIBUYA': 'SHIBUYA',
    'SIDEL': 'SIDEL',
    'BMB': 'BMB',
    'SIPA': 'SIPA',
    'ENGEL': 'ENGEL',
    'NETSTAL': 'NETSTAL',
    'SACMI': 'SACMI',
    'HUAYAN': 'HUAYAN',
    'DEMAG': 'Sumitomo (SHI) Demag',
    'SUMITOMO': 'Sumitomo (SHI) Demag',
    'SIAPI':'SIAPI',
    'Nissei':'Nissei ASB',
    'ASB':'Nissei ASB',
    'SAPI':'SIAPI'
}

def apply_priority(name):
    upper_name = str(name).upper()
    for keyword, result in priority_brands.items():
        if keyword in upper_name:
            return result
    return None  # No override

def cluster_names(df, entity_col, cleaned_col, brand_col, final_col):
    df[cleaned_col] = df[entity_col].fillna('').apply(clean_name)
    df[brand_col] = df[cleaned_col].apply(extract_brand)

    results = []
    for brand, group in df.groupby(brand_col):
        group = group.sort_values(cleaned_col).reset_index(drop=True)
        master_list = []
        cleaned_names = []

        for index, row in group.iterrows():
            current_original = row[entity_col]
            current_clean = row[cleaned_col]

            # Priority rule check
            priority_match = apply_priority(current_original)
            if priority_match:
                cleaned_names.append(priority_match)
                continue

            found_match = False
            best_score = 0
            best_master_original = None

            for master_original, master_clean in master_list:
                score = fuzz.token_set_ratio(master_clean, current_clean)
                if score > 80 and score > best_score:
                    best_score = score
                    best_master_original = master_original
                    found_match = True

            if found_match:
                cleaned_names.append(best_master_original)
            else:
                cleaned_names.append(current_original)
                master_list.append((current_original, current_clean))

        group[final_col] = cleaned_names
        results.append(group)

    return pd.concat(results, ignore_index=True)

# === Auto header detection ===
def detect_header_row(file, sheet_name, required_columns):
    for header_row in range(10):
        try:
            df_sample = pd.read_excel(file, sheet_name=sheet_name, header=header_row, nrows=1)
            if all(col in df_sample.columns for col in required_columns):
                return header_row
        except Exception:
            continue
    return None

# === Upload file ===
uploaded_file = st.file_uploader("📂 Upload Excel file", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    sheet_name = st.selectbox("📑 Select sheet to process:", xls.sheet_names)

    required_columns = ['Buyer', 'Supplier']
    header_row = detect_header_row(uploaded_file, sheet_name, required_columns)

    if header_row is None:
        st.error("❌ Could not find both 'Buyer' and 'Supplier' columns in the first 10 rows.")
        st.stop()

    # Read full sheet with detected header
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row)

    # Step 1: Cluster Buyer names
    st.write("🔄 Clustering Buyer names...")
    df = cluster_names(df, 'Buyer', 'Buyer Cleaned', 'Buyer Brand', 'Buyer Cleaned Final')

    # Step 2: Cluster Supplier names
    st.write("🔄 Clustering Supplier names...")
    df = cluster_names(df, 'Supplier', 'Supplier Cleaned', 'Supplier Brand', 'Supplier Cleaned Final')

    # Step 3: Reorder & drop temp columns
    cols = df.columns.tolist()

    if 'Buyer' in cols and 'Buyer Cleaned Final' in cols:
        cols.insert(cols.index('Buyer') + 1, cols.pop(cols.index('Buyer Cleaned Final')))
    if 'Supplier' in cols and 'Supplier Cleaned Final' in cols:
        cols.insert(cols.index('Supplier') + 1, cols.pop(cols.index('Supplier Cleaned Final')))

    # Remove temp/generated columns
    for col_to_drop in ['Buyer Cleaned', 'Supplier Cleaned', 'Buyer Brand', 'Supplier Brand']:
        if col_to_drop in df.columns:
            df.drop(columns=col_to_drop, inplace=True)

    df = df[[col for col in cols if col in df.columns]]

    # Final sort
    if 'Buyer Cleaned Final' in df.columns and 'Supplier Cleaned Final' in df.columns:
        df = df.sort_values(['Buyer Cleaned Final', 'Supplier Cleaned Final'])

    # === Write to Excel with highlight ===
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Clustered')
        workbook = writer.book
        worksheet = writer.sheets['Clustered']

        # Highlight "Buyer Cleaned Final" and "Supplier Cleaned Final"
        highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        for col_idx, col in enumerate(df.columns, start=1):
            if col in ["Buyer Cleaned Final", "Supplier Cleaned Final"]:
                worksheet.cell(row=1, column=col_idx).fill = highlight_fill

    output.seek(0)

    # === Download button ===
    st.success("✅ Clustering complete!")
    st.download_button(
        label="⬇️ Download Cleaned Excel",
        data=output.getvalue(),
        file_name="cleaned_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
