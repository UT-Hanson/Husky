import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import difflib


st.set_page_config(page_title="🧠 IMM Machine Model Extractor", layout="wide")
st.title("🏭 IMM Machine Model & Buyer Application Extractor")

# === Upload files ===
patterns_file = st.file_uploader("📥 Upload regex pattern Excel (regex.xlsx)", type=["xlsx"])
input_file = st.file_uploader("📄 Upload input Excel file", type=["xlsx"])
buyer_app_file = st.file_uploader("🔍 Upload Buyer → Application match file", type=["xlsx"])
region_file = st.file_uploader("🌍 Upload Country → Region match file", type=["xlsx"])


# === Helper: load with sheet selection ===
def load_excel_with_sheet_selector(file, label):
    if not file:
        return None
    xls = pd.ExcelFile(file)
    sheet = st.selectbox(f"📑 Select sheet for {label}", xls.sheet_names, key=label)
    return pd.read_excel(xls, sheet_name=sheet)


# === Load all files ===
patterns_df = load_excel_with_sheet_selector(patterns_file, "Regex Pattern File")
input_df = load_excel_with_sheet_selector(input_file, "Input File")
buyer_df = load_excel_with_sheet_selector(buyer_app_file, "Buyer Application File")

if patterns_df is not None and input_df is not None:
    try:
        patterns_df.columns = patterns_df.columns.str.strip()
        machine_patterns = patterns_df['Pattern'].dropna().tolist()
        df = input_df.copy()

        if "Tonnage" not in df.columns:
            df.insert(0, "Tonnage", "")
        if "Application" not in df.columns:
            df.insert(0, "Application", "")
        # === Load buyer application map ===
        buyer_map = {}
        if buyer_df is not None:
            buyer_df.columns = buyer_df.columns.str.strip()
            buyer_map = dict(zip(buyer_df['Buyer'].str.upper().str.strip(), buyer_df['Buyer Potential Application']))


        # === Model extraction ===
        def extract_model(description):
            text = str(description).upper().replace("  ", " ")
            for pattern in machine_patterns:
                try:
                    match = re.search(pattern, text)
                except re.error:
                    continue
                if match:
                    groups = match.groups()
                    if not groups:
                        continue
                    series = groups[0].strip().title()
                    parts = [g.strip() for g in groups[1:] if g and g.strip()]
                    if not parts:
                        return series
                    if len(parts) == 1:
                        part = parts[0]
                        if "/" in part:
                            return f"{series} {'-'.join(part.split('/'))}"
                        split_parts = re.findall(r'\d+[A-Z]?|\d{3,5}', part)
                        return f"{series} {'-'.join(split_parts)}" if split_parts else f"{series} {part}"
                    else:
                        cleaned_parts = [re.sub(r"\s+", "", p) for p in parts]
                        return f"{series} {'-'.join(cleaned_parts)}"
            return ""


        def extract_model_series(model):
            model = str(model).strip()
            if not model:
                return ""

            # Match characters before the first number (optional space or hyphen included)
            match = re.match(r'^([A-Za-z\- ]+)', model)
            if match:
                series = match.group(1).strip().replace("-", " ")
                return re.sub(r'\s+', ' ', series).title()
            return ""





        # === Tonnage extraction ===
        def extract_tonnage(row):
            import re
            model = str(row.get("Model", "")).upper()
            supplier = str(row.get("Supplier", "")).upper()
            if not model:
                return ""
            try:
                if "NETSTAL" in supplier:
                    m = re.search(r'(\d+)-\d+', model)
                    return round(int(m.group(1)) / 10, 1) if m else ""
                elif any(s in supplier for s in ["DEMAG", "SUMITOMO", "UBE"]):
                    m = re.search(r'(\d{2,5})', model)
                    return int(m.group(1)) if m else ""
                elif "ENGEL" in supplier:
                    numbers = re.findall(r'\d{2,5}', model)
                    if len(numbers) >= 2:
                        return round(int(numbers[1]), 1)
                elif "ARBURG" in supplier:
                    m = re.findall(r'\d{3,5}', model)
                    return round(int(m[1]) / 10, 1) if len(m) >= 2 else ""
                elif "BMB" in supplier:
                    m = re.search(r'(\d{2,5})', model)
                    return int(m.group(1)) * 10 if m else ""
                elif any(s in supplier for s in ["AOKI", "NAIGAI"]):
                    m = re.search(r'AL[-\s]?[\dA-Z]+[-\s]?(\d{2,4})', model)
                    return int(m.group(1)) * 10 if m else ""
                elif any(s in supplier for s in ["ASB", "NISSEI"]):
                    m = re.search(r'ASB[-\s]?(\d{2,4})', model)
                    if m:
                        return int(m.group(1))
                    m = re.search(r'(\d{2,4})', model)
                    return int(m.group(1)) if m else ""

                # === New: HUAYAN and SACMI
                elif any(s in supplier for s in ["HUAYAN", "SACMI"]):
                    m = re.search(r'[-\s]?(\d{2,4})', model)
                    return int(m.group(1)) if m else ""

                # === New: SIPA XFORM and XTREME
                elif "SIPA" in supplier or "XFORM" in model or "XTREME" in model:
                    if "XFORM" in model:
                        m = re.search(r'XFORM[-\s]?(\d{3,4})', model)
                        return int(m.group(1)) if m else ""
                    elif "XTREME" in model:
                        return ""  # XTREME is cavity-based, no tonnage

            except Exception:
                return ""
            return ""


        def classify_tonnage_range(tonnage_str):
            try:
                if pd.isna(tonnage_str):
                    return ""
                tonnage = float(str(tonnage_str).replace(",", ""))
                if tonnage < 300:
                    return "Small (<300)"
                elif 300 <= tonnage <= 799:
                    return "Medium (300–799)"
                elif tonnage >= 800:
                    return "Large (800+)"
            except:
                return ""


        # === Buyer Application Matching ===
        def match_buyer_app(row):
            buyer = str(row.get("Buyer Cleaned Final", "")).strip().upper()
            supplier = str(row.get("Supplier Cleaned Final", "")).strip().upper()

            if buyer in buyer_map:
                return buyer_map[buyer]

            # Fallback keyword rules
            elif "TRAD" in buyer:
                return "Trading"
            elif "AUTO" in buyer:
                return "Automotive"
            elif "LOGIST" in buyer:
                return "Logistic"
            elif "PACK" in buyer:
                return "Packaging"
            elif "ELECTR" in buyer:
                return "Electronic"
            elif buyer[:4] == supplier[:4] and buyer:
                return supplier
            else:
                return ""


        if region_file and "Buyer Country" in df.columns:
            region_df = load_excel_with_sheet_selector(region_file, "Region File")
            if region_df is not None:
                region_df.columns = region_df.columns.str.strip()
                region_df = region_df[['Country', 'Region']].dropna()


                # Normalize function
                def normalize(text):
                    return re.sub(r'\s+', ' ', str(text).strip().lower())


                # Create region lookup dict
                region_dict = {normalize(country): region for country, region in
                               zip(region_df['Country'], region_df['Region'])}


                # Fuzzy match function with priority logic
                def fuzzy_match_region(buyer_country):
                    if pd.isna(buyer_country):
                        return ""

                    norm_buyer = normalize(buyer_country)

                    # === Priority keyword rules ===
                    upper_buyer = str(buyer_country).upper()
                    if "KINGDOM" in upper_buyer:
                        return "West Europe"
                    if "OF AM" in upper_buyer or "UNITED STATES" in upper_buyer:
                        return "North America"
                    if "EMIRATES" in upper_buyer:
                        return "Middle East"

                    # Fuzzy match
                    matches = difflib.get_close_matches(norm_buyer, region_dict.keys(), n=1, cutoff=0.8)
                    if matches:
                        return region_dict[matches[0]]

                    # Match first 5 characters
                    for key in region_dict:
                        if norm_buyer[:5] == key[:5]:
                            return region_dict[key]

                    return ""


                df["Buyer Region"] = df["Buyer Country"].apply(fuzzy_match_region)

        if "Product Description" not in df.columns:
            st.error("❌ 'Product Description' column not found.")
            st.stop()

        df["Tonnage Range"] = df["Tonnage"].apply(classify_tonnage_range)
        df["Model"] = df["Product Description"].apply(extract_model)
        df["Model Series"] = df["Model"].apply(extract_model_series)
        df["Tonnage"] = df.apply(extract_tonnage, axis=1)

        if "Model Series" in df.columns and "Tonnage" in df.columns:
            ms_index = df.columns.get_loc("Model Series")
            reordered_cols = df.columns.tolist()
            reordered_cols.insert(ms_index + 1, reordered_cols.pop(reordered_cols.index("Tonnage")))
            df = df[reordered_cols]


        # Add Buyer Potential Application



        def classify_product_type(row):
            supplier = str(row.get("Supplier", "")).upper()
            model = str(row.get("Model", "")).upper().strip()

            if any(x in supplier for x in ["BMB", "ARBURG", "DEMAG", "ENGEL"]):
                return "Packaging"
            elif any(x in supplier for x in ["HUAYAN", "SIPA", "SACMI"]):
                return "PET"
            elif "NETSTAL" in supplier:
                if not model:
                    return ""
                return "PET" if "PET" in model else "Packaging"
            else:
                return ""


        def assign_application_sub_category(row):
            product_type = str(row.get("Product Type Focus (Packaging/PET)", "")).strip().upper()
            potential_app = str(row.get("Buyer Potential Application", "")).strip().upper()

            if product_type == "PET":
                return "PET"
            if potential_app == "PACKAGING":
                return "Thin Wall Packaging and Pails"
            if "CLOSURE" in potential_app:
                return "Closure"
            if any(k in potential_app for k in ["ENGEL", "ARBURG", "NETSTAL"]):
                return "Inter Company Shipment"
            if any(k in potential_app for k in ["MEDICAL", "HEALTHCARE", "PHARMACEUTICAL", "LAB CONSUMABLE"]):
                return "Medical, Pharma, and Lab consumable"
            return "Other"


        if "Buyer" in df.columns and "Supplier" in df.columns:
            df["Buyer Potential Application"] = df.apply(match_buyer_app, axis=1)
            # Fix blank Product Type Focus BEFORE calling assign_application_sub_category
            # 1. Create Product Type column first
            df["Product Type Focus (Packaging/PET)"] = df.apply(classify_product_type, axis=1)

            # 2. Fix blank values (force PET if blank)
            df["Product Type Focus (Packaging/PET)"] = df["Product Type Focus (Packaging/PET)"].fillna("").replace("",
                                                                                                                   "PET")

            # 3. Now it's safe to calculate Application Sub Category
            df["Application Sub Category"] = df.apply(assign_application_sub_category, axis=1)

        if "Product Type Focus (Packaging/PET)" in df.columns and "Application Sub Category" in df.columns:
            pt_index = df.columns.get_loc("Product Type Focus (Packaging/PET)")
            reordered_cols = df.columns.tolist()
            reordered_cols.insert(pt_index + 1, reordered_cols.pop(reordered_cols.index("Application Sub Category")))
            df = df[reordered_cols]

        # Reorder columns
        # === Reorder Columns ===
        final_order = [
            "Buyer", "Buyer Cleaned Final", "Buyer Potential Application", "Product Type Focus (Packaging/PET)",
            "Application Sub Category",
            "Buyer Country", "Buyer Region", "Supplier", "Supplier Cleaned Final", "Supplier Country",
            "HS Code", "Product Description", "Product", "Model", "Model Series", "Tonnage", "Tonnage Range",
            "Quantity", "Application", "Unit", "Value", "Trade Direction", "Date", "Year", "Data Source"
        ]

        # Keep any extra columns not listed above
        remaining = [col for col in df.columns if col not in final_order]
        df = df[[col for col in final_order if col in df.columns] + remaining]

        # === Columns to highlight and format ===
        highlight_cols = [
            "Buyer Potential Application", "Product Type Focus (Packaging/PET)",
            "Model", "Model Series", "Tonnage", "Tonnage Range"
        ]
        numeric_cols = ["Tonnage", "Quantity", "Value", "Tonnage Range"]

        # === Save clean (numeric) copy for export ===
        styled_df = df.copy()

        # === Write DataFrame to BytesIO with openpyxl ===
        output = BytesIO()
        styled_df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        # === Load workbook and apply styling ===
        wb = load_workbook(output)
        ws = wb.active

        highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow
        number_format = '#,##0'

        # Get header index mapping
        header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        # === Highlight ONLY header cells ===
        for col_name in highlight_cols:
            if col_name in header:
                col_idx = header[col_name]
                ws.cell(row=1, column=col_idx).fill = highlight_fill

        # === Format number cells in data rows ===
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_name in numeric_cols:
                if col_name in header:
                    cell = row[header[col_name] - 1]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format

        # Save updated workbook to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ✅ Download Button
        st.download_button(
            label="⬇️ Download Processed File",
            data=output,
            file_name="processed_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ An error occurred: {e}")

