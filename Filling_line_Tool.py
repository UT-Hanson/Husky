import streamlit as st
import pandas as pd
import re
import difflib
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment

st.set_page_config(page_title="🧠 Product Line Detector", layout="wide")
st.title("🔍 Detect Product and Product Line from Description")

# === Upload files ===
data_file = st.file_uploader("📥 Upload Data File", type=["xlsx", "csv"])
match_file = st.file_uploader("📄 Upload Match File (Pattern to Lines)", type=["xlsx", "csv"])
region_file = st.file_uploader("🌍 Upload Region File (Country to Region)", type=["xlsx", "csv"])

# === Sheet selection helper ===
def load_sheet(file, label):
    if not file or file.name.endswith(".csv"):
        return pd.read_csv(file) if file else None

    xls = pd.ExcelFile(file)
    sheet_name = st.selectbox(f"📑 Select sheet for {label}", xls.sheet_names, key=label)
    return pd.read_excel(xls, sheet_name=sheet_name)

# === Load files with sheet selection ===
data_df = load_sheet(data_file, "Data File")
match_df = load_sheet(match_file, "Match File")
region_df = load_sheet(region_file, "Region File") if region_file else None


# === Application Keywords ===
# Lowercased version of application keywords for reliable matching
application_keywords_raw = [
    'water', 'milk', 'beer', 'juice', 'soft drink', 'carbonated', 'tea', 'coffee', 'energy drink',
    'wine', 'soda', 'syrup', 'yogurt', 'liquid', 'beverage', 'dairy', 'cocktail', 'liqueur',
    'mineral', 'spring', 'flavored', 'seltzer',
    'молока', 'вода', 'сок', 'пиво'
    , 'напиток', 'газированный','ГАЗИРОВАННЫХ',"НАПИТКОВ "
    'agua', 'leche', 'cerveza', 'zumo', 'bebida', 'gaseosa',
    'süt', 'bira', 'meyve suyu', 'içecek',
    'paani', 'doodh', 'juice', 'sharab', 'cold drink'
]
application_keywords = [kw.lower() for kw in application_keywords_raw]

def detect_application(text):
    text = str(text).lower()
    for keyword in application_keywords:
        if keyword in text:
            return keyword
    return ""


def classify_product_line(row):
    value = row.get('Value', 0)
    line = str(row.get('Product Line', '')).lower()

    if value < 200000:
        return "Other"
    if any(x in line for x in ['pet', 'aseptic']):
        return "FILLING LINE - PET"
    elif any(x in line for x in ['glass']):
        return "FILLING LINE - Glass"
    elif any(x in line for x in ['can']):
        return "FILLING LINE - Can"
    elif any(x in line for x in ['keg']):
        return "FILLING LINE - Keg"
    elif 'fill' in line:
        return "FILLING LINE - Unspecified"

    return "Other"


if data_file and match_file:
    if data_file.name.endswith(".csv"):
        data_df = pd.read_csv(data_file)
    else:
        data_df = pd.read_excel(data_file)

    if match_file.name.endswith(".csv"):
        match_df = pd.read_csv(match_file)
    else:
        match_df = pd.read_excel(match_file)

    data_df.columns = data_df.columns.str.strip()
    match_df.columns = match_df.columns.str.strip()

    if region_file and "Buyer Country" in data_df.columns:
        if region_file.name.endswith(".csv"):
            region_df = pd.read_csv(region_file)
        else:
            region_df = pd.read_excel(region_file)

        region_df.columns = region_df.columns.str.strip()
        region_df = region_df[['Country', 'Region']].dropna()


        def normalize(text):
            return re.sub(r'\s+', ' ', str(text).strip().lower())


        region_dict = {normalize(c): r for c, r in zip(region_df['Country'], region_df['Region'])}


        def fuzzy_region_match(buyer_country):
            if pd.isna(buyer_country):
                return ""

            upper_buyer = str(buyer_country).upper()

            # === Priority keyword rules ===
            if "KINGDOM" in upper_buyer:
                return "West Europe"
            if "OF AM" in upper_buyer or "UNITED STATES" in upper_buyer:
                return "North America"
            if "EMIRATES" in upper_buyer:
                return "Middle East"

            norm_buying = normalize(buyer_country)

            # Fuzzy match (80%)
            matches = difflib.get_close_matches(norm_buying, region_dict.keys(), n=1, cutoff=0.8)
            if matches:
                return region_dict[matches[0]]

            # Try first 5-letter match
            for key in region_dict:
                if norm_buying[:5] == key[:5]:
                    return region_dict[key]

            return ""


        data_df["Buyer Region"] = data_df["Buyer Country"].apply(fuzzy_region_match)

    # === Product Matching ===
    data_df["Product"] = ""
    data_df["Product Line"] = ""

    for _, row in match_df.iterrows():
        pattern = str(row["Pattern"]).strip().upper()
        line = str(row["Lines"]).strip()
        mask = (data_df["Product Line"] == "") & data_df["Product Description"].str.upper().str.contains(pattern, na=False)
        data_df.loc[mask, "Product"] = pattern
        data_df.loc[mask, "Product Line"] = line

    # === Application and Classification ===
    data_df["Application"] = data_df["Product Description"].apply(detect_application)
    data_df["Product Line Classification"] = data_df.apply(classify_product_line, axis=1)

    # === Reorder Columns ===
    cols = list(data_df.columns)
    if "Product Description" in cols:
        idx = cols.index("Product Description")
        new_cols = ["Product", "Product Line", "Application", "Product Line Classification"]
        for col in new_cols:
            if col in cols:
                cols.remove(col)
        cols = cols[:idx + 1] + new_cols + cols[idx + 1:]
        if "Buyer Country" in cols:
            buyer_idx = cols.index("Buyer Country")
            if "Buyer Region" in cols:
                cols.remove("Buyer Region")
                cols.insert(buyer_idx + 1, "Buyer Region")
        data_df = data_df[cols]

    # === Display preview ===
    st.success("✅ Matching complete! Preview below:")
    st.dataframe(data_df)

    # === Excel Export with Highlight and Formatting ===
    def to_excel_with_header_highlight(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Result", startrow=3)
            wb = writer.book
            ws = writer.sheets["Result"]

            # Add multi-cell description
            #intro = ("This file classifies product lines based on value and product line name. "
            #         "If Value < 200,000 → 'Other'. If Product Line includes 'PET' or 'Aseptic' → 'PET'. "
               #      "If it includes 'glass', 'can', or 'keg' → 'Glass / Can / Keg'. "
            #         "If it only contains 'fill' → 'Unspecified'. Else → 'Other'.")
           # ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=6)
            #cell = ws.cell(row=1, column=1)
           # cell.value = intro
           # cell.alignment = Alignment(wrap_text=True, vertical="top")
           # cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # Highlight selected columns
            highlight_cols = ["Product", "Product Line", "Buyer Region","Application", "Product Line Classification"]
            highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name in highlight_cols:
                    ws.cell(row=4, column=col_idx).fill = highlight_fill

            # Format numeric columns
            numeric_cols = ["Quantity", "Value"]
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name in numeric_cols:
                    for row in range(5, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'

        output.seek(0)
        return output.getvalue()

    excel_data = to_excel_with_header_highlight(data_df)
    st.download_button(
        label="📥 Download Result as Excel (Highlighted)",
        data=excel_data,
        file_name="processed_product_lines_highlighted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
